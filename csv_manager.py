"""
Módulo para gerenciar operações com arquivos Excel/CSV
"""
import streamlit as st
import pandas as pd
import os
from typing import Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

class CSVManager:
    """
    Classe para gerenciar operações de leitura e escrita de arquivos Excel
    """
    
    DEFAULT_FILENAME = 'vagas_extraidas.xlsx'
    
    def __init__(self, filename: str = DEFAULT_FILENAME):
        self.filename = filename
    
    def _create_excel(self, df: pd.DataFrame) -> Workbook:
        """Cria um Workbook Excel formatado a partir de um DataFrame."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vagas"

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', start_color='2E4057')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align = Alignment(vertical='center', wrap_text=False)
        link_align = Alignment(vertical='center', wrap_text=False)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_names = {
            'empresa': 'Empresa',
            'cidade': 'Cidade',
            'estado': 'Estado',
            'data_publicacao': 'Data de Publicação',
            'url': 'URL da Vaga',
        }

        columns = [c for c in col_names if c in df.columns]
        headers = [col_names[c] for c in columns]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 30

        fill_alt = PatternFill('solid', start_color='F0F4F8')
        fill_normal = PatternFill('solid', start_color='FFFFFF')

        for row_idx, row in enumerate(df[columns].itertuples(index=False), start=2):
            fill = fill_alt if row_idx % 2 == 0 else fill_normal
            for col_idx, value in enumerate(row, start=1):
                col_name = columns[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.fill = fill
                cell.border = border
                if col_name == 'url' and value and value != 'N/A':
                    cell.hyperlink = value
                    cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')
                    cell.alignment = link_align
                else:
                    cell.alignment = cell_align
            ws.row_dimensions[row_idx].height = 18

        col_widths = {
            'empresa': 35,
            'cidade': 22,
            'estado': 10,
            'data_publicacao': 20,
            'url': 55,
        }
        for col_idx, col_name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        return wb

    def save_dataframe(self, df: pd.DataFrame, filename: Optional[str] = None) -> Tuple[bool, str]:
        """Salva um DataFrame em arquivo Excel formatado."""
        try:
            if df is None or df.empty:
                return False, "❌ DataFrame vazio. Nada para salvar."
            
            save_file = filename or self.filename
            wb = self._create_excel(df)
            wb.save(save_file)
            return True, f"✅ Dados salvos com sucesso em '{save_file}'!"
            
        except Exception as e:
            return False, f"❌ Erro ao salvar arquivo: {str(e)}"
    
    def load_dataframe(self, filename: Optional[str] = None) -> Tuple[Optional[pd.DataFrame], str]:
        """Carrega um DataFrame de um arquivo Excel."""
        try:
            load_file = filename or self.filename
            
            if not os.path.exists(load_file):
                csv_file = load_file.replace('.xlsx', '.csv')
                if os.path.exists(csv_file):
                    df = pd.read_csv(csv_file)
                else:
                    return None, f"❌ Arquivo '{load_file}' não encontrado."
            else:
                df = pd.read_excel(load_file)
            
            if df.empty:
                return None, f"⚠️ Arquivo está vazio."
            
            return df, f"✅ Carregadas {len(df)} vagas!"
            
        except Exception as e:
            return None, f"❌ Erro ao carregar arquivo: {str(e)}"
    
    def file_exists(self, filename: Optional[str] = None) -> bool:
        """Verifica se o arquivo existe."""
        check_file = filename or self.filename
        return os.path.exists(check_file) or os.path.exists(check_file.replace('.xlsx', '.csv'))
    
    def delete_file(self, filename: Optional[str] = None) -> Tuple[bool, str]:
        """Deleta o arquivo."""
        try:
            delete_file = filename or self.filename
            if os.path.exists(delete_file):
                os.remove(delete_file)
                return True, f"✅ Arquivo deletado com sucesso!"
            csv_file = delete_file.replace('.xlsx', '.csv')
            if os.path.exists(csv_file):
                os.remove(csv_file)
                return True, f"✅ Arquivo deletado com sucesso!"
            return False, f"❌ Arquivo não encontrado."
        except Exception as e:
            return False, f"❌ Erro ao deletar arquivo: {str(e)}"
    
    def get_download_data(self, df: pd.DataFrame) -> Tuple[bytes, str]:
        """Prepara os dados para download como Excel."""
        if df is None or df.empty:
            return b"", "dados_vazios.xlsx"
        
        wb = self._create_excel(df)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"vagas_maringa_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return buffer.getvalue(), filename


def create_csv_manager_widgets():
    csv_manager = CSVManager()
    file_exists = csv_manager.file_exists()
    
    if file_exists:
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📊 Carregar Excel salvo", use_container_width=True):
                df, message = csv_manager.load_dataframe()
                if df is not None:
                    if 'título' in df.columns:
                        df = df.drop('título', axis=1)
                    if 'titulo' in df.columns:
                        df = df.drop('titulo', axis=1)
                    st.session_state['loaded_df'] = df
                    st.session_state['df'] = df
                    st.success(message)
                    st.rerun()
                else:
                    st.error(message)
        
        with col2:
            if st.button("🗑️ Deletar arquivo", use_container_width=True):
                success, message = csv_manager.delete_file()
                if success:
                    st.success(message)
                    if 'loaded_df' in st.session_state:
                        st.session_state['loaded_df'] = None
                    if 'df' in st.session_state:
                        st.session_state['df'] = None
                    st.rerun()
                else:
                    st.error(message)
    else:
        st.info("ℹ️ Nenhum arquivo salvo encontrado.")
    
    return csv_manager


def save_current_dataframe(df: pd.DataFrame) -> Tuple[bool, str]:
    csv_manager = CSVManager()
    return csv_manager.save_dataframe(df)
